import datetime
import os
import time
from datetime import timedelta
from time import sleep

import telebot
import gspread
import win32security
import shutil
import pythoncom
import openpyxl
import win32com.client
import requests
import json
import threading

from tqdm import tqdm
from ftplib import FTP
from dates import *
from PIL import Image
from threading import Thread

# Класс времён
class times:
    # Время сейчас
    today = datetime.datetime.today()
    todaytime = today.strftime("%H:%M:%S")
    # Время сканирования папок с фотографиями
    #timetoScan = today.time().strftime("%H:%M")
    timetoScan = (today + datetime.timedelta(minutes=15)).strftime("%H:%M")
    # Время для работы изменения Call-центра
    timetoChangeCallCenter = (today + datetime.timedelta(minutes=2)).strftime("%H:%M")
    # Время для сбора статистики по звонкам
    #timetoCollectionOfInformation = datetime.time(0, 5).strftime("%H:%M")
    timetoCollectionOfInformation = (today + datetime.timedelta(minutes=3)).strftime("%H:%M")
    # Время собрания (пока не используется)
    #timetoOffCallCenterOnMeeting = datetime.time(16, 0).strftime("%H:%M")
    # Время сбора статистики по недельной работе прикрепления фотографий к карточкам товаров
    #timetoGenerationStatUploadPhotos = datetime.time(2, 30).strftime("%H:%M")
    timetoGenerationStatUploadPhotos = (today + datetime.timedelta(minutes=4)).strftime("%H:%M")
    #timetoGenerationStatUploadPhotos = today.time().strftime("%H:%M")
    # Время для проверки 2.0 на сканирование фотографий
    timetoScan_2_0 = datetime.time(3, 47).strftime("%H:%M")
    #timetoScan_2_0 = today.time().strftime("%H:%M")
    # Время для проверки файлов на обновления
    timetoScanUpdateFiles = today.time().strftime("%H:%M")
    #timetoScanUpdateFiles = (today + datetime.timedelta(minutes=1)).strftime("%H:%M")

    def nexttime(self, argument):
        match(argument):
            # Время сканирования папок с фотографиями
            case "TimeToScan":
                nexthour = datetime.datetime.today().hour + 1
                if nexthour == 24:
                    nexthour = 0
                self.timetoScan = datetime.time(nexthour, 0).strftime("%H:%M")
                return self.timetoScan

            # Время для сбора статистики по звонкам
            case "timetoCollectionOfInformation":
                times.timetoCollectionOfInformation = datetime.time(0, 5).strftime("%H:%M")
                return times.timetoCollectionOfInformation

            # Время для работы изменения Call-центра
            case "timetoChangeCallCenter":
                nexthour = datetime.datetime.today().hour + 1
                if nexthour == 24:
                    nexthour = 0
                times.timetoChangeCallCenter = datetime.time(nexthour, 10).strftime("%H:%M")
                return times.timetoChangeCallCenter

            # Время сбора статистики по недельной работе прикрепления фотографий к карточкам товаров
            case "timetoGenerationStatUploadPhotos":
                times.timetoGenerationStatUploadPhotos = datetime.time(2, 30).strftime("%H:%M")
                return times.timetoGenerationStatUploadPhotos

            # Время для проверки файлов на обновления
            case "timetoScanUpdateFiles":
                times.timetoScanUpdateFiles = datetime.time(4, 0).strftime("%H:%M")
                return times.timetoScanUpdateFiles

            case _:
                pass

# Класс работы с фотографиями
class class_photos(object):
    # Массив локальных файлов
    masslocal = [[], [], [], [], []]
    # Массив размеров локальных файлов
    masslocalsize = [[], [], [], [], []]
    # Массив удалённых файлов
    massremote = [[], [], [], [], []]
    # Массив дат создания удалённых файлов
    massremotesizes = [[], [], [], [], []]
    # Массив загружаемых фотографий
    massnewphotos = [0, 0, 0, 0, 0]

    def __init__(self, argument, timetowaitingfunction):
        self.date = argument
        self.timetowaitingfunction = timetowaitingfunction

    # Функция последовательного запуска функций
    def startprocessing(self):
        # Функция разбора фотографий
        self.scanfolderwithimages()

        # Необходима конструкция, которая будет обрабатывать недоступность папок как локальных, так и удалённых.
        # Вызов функции сканирования локальных папок
        self.scanfolderforimages()
        #thr1 = threading.Thread(target = self.scanfolderforimages)
        #thr1.start()
        self.scanfilesinremoteserver()
        # Вызов функции сканирования удалённых папок
        #thr2 = threading.Thread(target = self.scanfilesinremoteserver)
        #thr2.start()

        # Ожидание окончания потоков
        #thr1.join(self.timetowaitingfunction)
        #thr2.join(self.timetowaitingfunction)

        # Запуск таймера по ограничению работы потоков
        #thr3 = threading.Thread(target = self.killingthreads, args=(timetowaitingfunction, thr1, thr2))
        #thr3.start()

        # Вызов функции выявления различия файлов на локальном и удалённом сервере
        self.comparisonlists()

    # Функция подгонки фотографий под единый формат
    def changefromat(self, folder):
        print("Функция подгонки фотографий под единый формат папки", folder)
        print("==========================")
        #print(folder)
        #print(len(os.listdir(folder)))
        #print(os.listdir(folder))
        for element in tqdm(os.listdir(folder)):
            oldname = folder + element
            newname = folder + element.lower()
            #print(f"Название элемента [{oldname}]\t[{newname}]")
            #print(oldname.islower())
            if os.path.exists(oldname) == True and oldname.islower():
                os.rename(oldname, newname)
        print("==========================")

    # Функция ограничения потоков по времени
    def killingthreads(self, timelimit, thr1, thr2):

        #Таймер отсчёта времени оставшегося жить потокам
        while timelimit:
            timelimit -= 1
            time.sleep(1)
        # Убиваем потоки, если превышен лимит времени выполнения функции
        if thr1.is_alive():
            print("Поток сканирования локальных папок работает слишком долго. Убиваем его!")
            thr1._stop.set()
        elif thr2.is_alive():
            print("Поток сканирования удалённых папок работает слишком долго. Убиваем его!")
            thr2._stop.set()
        else:
            print("Потоки сканирования папок отработали в штатном режиме.")

    # Функция записи логов папок фотографий
    def createnewarrowinlogs(self, lenphotos):
        try:
            # Подключаемся к сервисному аккаунту
            gc = gspread.service_account(CREDENTIALS_FILE)
            # Подключаемся к таблице по ключу таблицы
            table = gc.open_by_key(sheetkey)
            # Открываем нужный лист
            worksheet = table.worksheet("LogsPhotos")
            # Получаем данные с листа
            dates = worksheet.get_values()
            # Получаем номер самой последней строки
            newstr = len(worksheet.col_values(1)) + 1
            # Вычисляем номер строки
            newnumber = newstr - 1
            # Определяем время выполения операции
            today = datetime.datetime.today().strftime("%d.%m.%Y | %H:%M:%S")
            # Добавляем строку в конец фаила логгирования
            worksheet.update_cell(newstr, 1, newnumber)
            worksheet.update_cell(newstr, 2, today)
            worksheet.update_cell(newstr, 3, lenphotos)
        except:
            print("Логгирование фотографий сломалось(")

    # Функция сбора статистики по загруженным фотографиям
    def statisticsphotos(self, pathimage, massnewphotos):
        sd = win32security.GetFileSecurity(pathimage, win32security.OWNER_SECURITY_INFORMATION)
        owner_sid = sd.GetSecurityDescriptorOwner()
        match (str(owner_sid)):
            case masssotr.PySID_fleysner:
                massnewphotos[0] += 1
            case masssotr.PySID_kireev:
                massnewphotos[1] += 1
            case masssotr.PySID_pushkar:
                massnewphotos[2] += 1
            case masssotr.PySID_ivanov:
                massnewphotos[3] += 1
            case _:
                massnewphotos[4] += 1
        return massnewphotos

    # Функция получения размера изображения
    def get_size_format(self, b, factor=1024, suffix="B"):
        """
        Scale bytes to its proper byte format
        e.g:
            1253656 => '1.20MB'
            1253656678 => '1.17GB'
        """
        for unit in ["", "K", "M", "G", "T", "P", "E", "Z"]:
            if b < factor:
                return f"{b:.2f}{unit}{suffix}"
            b /= factor
        return f"{b:.2f}Y{suffix}"

    # Функция конвертации изображения (уменьшения веса и подгонка под заданные параметры)
    def convertimage(self, path):
        # Размеры изображения на выходе
        width = 1920
        height = 1440
        # Загружаем фотографию в память
        img = Image.open(path)
        # Первоначальный размер картинки
        olddimensions = img.size
        # Получаем размер изображения до компрессии
        image_size = os.path.getsize(path)
        oldsize = self.get_size_format(image_size)
        # Преобразуем изображение приводя его к нужным высоте и ширине и уменьшая размер
        img.thumbnail(size=(width, height))
        if img.height > 1080:
            difference_height = (height - 1080) / 2
            img = img.crop((0, 0 + difference_height, 1920, height - difference_height))
        # Сохраняем изображение
        img.save(path, optimize=True, quality=95)
        # Получаем новые размеры картинки
        newdimesions = img.size
        # Получаем размер изображение после компрессии
        image_size = os.path.getsize(path)
        newsize = self.get_size_format(image_size)
        # Печатаем в кносоль результат
        print(f"{path} с шириной, высотой: {olddimensions} и размером: {oldsize} была преобразована в: {newdimesions} и {newsize}")

    # Переименование и перемещение картинки по необходимому локальному пути
    def renameanduploadimage(self, pathimage, folder):
        lenmailfolder = 62
        # Начинаем с переименования картинки
        numberfolderfirst = str(pathimage)[lenmailfolder:]
        # Если папка четырёхзначная
        if numberfolderfirst[4] == "/":
            numberfoldersecond = str(numberfolderfirst)[:4]
        elif numberfolderfirst[3] == "/":
            numberfoldersecond = str(numberfolderfirst)[:3]
        else:
            numberfoldersecond = str(numberfolderfirst)[:5]
        # Название картинки
        namepic = numberfoldersecond + str(pathimage)[-4:]
        # Новый путь к картинке
        convertname = str(pathimage)[:lenmailfolder] + numberfoldersecond + "/" + namepic
        # Переименование картинки
        os.rename(pathimage, convertname)

        # Начинаем загрузку фотографии по необходимому местоположению
        newpathfile = str(pathimage)[:29] + str(folder) + "/" + namepic
        shutil.move(convertname, newpathfile)

    # Функция записи статистики по загруженным фотографиям в GoogleDocs
    def updatedatesuploadphotos(self, massnewphotos):
        try:
            # Подключаемся к сервисному аккаунту
            gc = gspread.service_account(CREDENTIALS_FILE)
            # Подключаемся к таблице по ключу таблицы
            table = gc.open_by_key(sheetkey)
            # Открываем нужный лист
            worksheet = table.worksheet("LogsPhotos")
            # Получаем данные из ячеек
            massolddates = [int(worksheet.get_values(masssotr.CellTable_fleysner)[0][0]),
                            int(worksheet.get_values(masssotr.CellTable_kireev)[0][0]),
                            int(worksheet.get_values(masssotr.CellTable_pushkar)[0][0]),
                            int(worksheet.get_values(masssotr.CellTable_ivanov)[0][0]),
                            int(worksheet.get_values(masssotr.CellTable_none)[0][0])]

            print(massolddates)
            # Новый массив для результата сложения
            newmass = []
            # Прибавляем новые значения к старым
            for elementfirst, elementsecond in zip(massolddates, massnewphotos):
                newmass += [elementfirst + elementsecond]
            print(newmass)
            # Обновляем значения в таблице
            worksheet.update_acell(masssotr.CellTable_fleysner, newmass[0])
            worksheet.update_acell(masssotr.CellTable_kireev, newmass[1])
            worksheet.update_acell(masssotr.CellTable_pushkar, newmass[2])
            worksheet.update_acell(masssotr.CellTable_ivanov, newmass[3])
            worksheet.update_acell(masssotr.CellTable_none, newmass[4])

        except Exception as e:
            print(f"Логгирование статистики фотографий сломалось: {e}")

    # Функция сканирование папки для разбора фотографий
    def scanfolderwithimages(self):
        # Массив загруженных фотографий
        massnewphotos = [0, 0, 0, 0, 0]
        # Получаем лист фаилов находящихся по адресу
        list = os.listdir(mainpathanalysis)
        # Проверка наличия фотографий
        # Если папок для разбора нет
        if list == []:
            print("\tДанные по импорту фотографий из папки для разбора отсутствуют")
        # Если есть папки для разбора
        else:
            # Выясняем количество папок
            lenphotos = len(list)
            # Логгирование фотографий
            self.createnewarrowinlogs(lenphotos)
            # Пробегаемся по элеметам
            for element in list:
                # Обыгрывание Thumbs.db
                if element == "Thumbs.db":
                    # Выясняем путь к этому фаилу
                    path = mainpathanalysis + "/" + element
                    # Пытаемся удалить данный фаил
                    try:
                        if os.access(path, os.R_OK and os.X_OK):
                            os.remove(path)
                    except PermissionError:
                        # Оператор заглушка равноценная отсутствию операции
                        pass
                # Если же это не файл Thumbs.db
                else:
                    # Выясняем путь к этому фаилу
                    pathfolder = mainpathanalysis + "/" + element
                    # Получаем данные о файлах по этому пути
                    nextlist = os.listdir(pathfolder)
                    # Если папка пуста, то пишем о пустой папке
                    if nextlist == []:
                        print("\t Папка ", element, " пуста")
                    # Если папка не пуста
                    else:
                        numberfolder = 1
                        for elem in tqdm(nextlist):
                            # Обыгрывание Thumbs.db
                            if elem == "Thumbs.db":
                                continue
                            else:
                                # Условие перебора количества фотографий, так как 6 папки нету
                                if numberfolder >= 6:
                                    break
                                else:
                                    # Выясняем путь к фаилу
                                    pathimage = pathfolder + "/" + elem
                                    # Функция сбора статистики по загруженным фотографиям
                                    massnewphotos = self.statisticsphotos(pathimage, massnewphotos)
                                    # Уменьшение веса и подгонка фотографии
                                    self.convertimage(pathimage)
                                    # Переименование и загрузка фотографии
                                    self.renameanduploadimage(pathimage, numberfolder)
                                    # Увеличиваем счётчик
                                    numberfolder = numberfolder + 1

            # После окончания загрузки фотографий по папкам удаляем папку
            for elem in list:
                # Обыгрывание Thumbs.db
                if element == "Thumbs.db":
                    path = mainpathanalysis + "/" + element
                    try:
                        if os.access(path, os.R_OK and os.X_OK):
                            os.remove(path)
                    except PermissionError:
                        pass
                else:
                    # Выясняем путь к папке
                    path = mainpathanalysis + "/" + elem
                    # Удаляем полностью папку
                    try:
                        shutil.rmtree(path)
                    except Exception as e:
                        print("Удаление невозможно. По причине ", e)
            t1 = Thread(target=self.updatedatesuploadphotos, args=(massnewphotos,))
            t1.start()
            print("Удаление папок завершено")

    # Функция сканирования локальных папок с фотографиями
    def scanfolderforimages(self):
        print(f"Начинаем сканирование данных из локальных папок")
        # Пробегаемся по массиву и заполняем данные по называнию файлов
        for element in os.listdir(mainpath):
            # Формируем путь к папкам
            pathfolder = mainpath + element + "/"
            match element:
                case "1":
                    #self.changefromat(pathfolder)
                    self.masslocal[0] = os.listdir(pathfolder)
                    if 'thumbs.db' in self.masslocal[0]:
                        self.masslocal[0].remove('thumbs.db')
                case "2":
                    #self.changefromat(pathfolder)
                    self.masslocal[1] = os.listdir(pathfolder)
                    if 'thumbs.db' in self.masslocal[1]:
                        self.masslocal[1].remove('thumbs.db')
                case "3":
                    #self.changefromat(pathfolder)
                    self.masslocal[2] = os.listdir(pathfolder)
                    if 'thumbs.db' in self.masslocal[2]:
                        self.masslocal[2].remove('thumbs.db')
                case "4":
                    #self.changefromat(pathfolder)
                    self.masslocal[3] = os.listdir(pathfolder)
                    if 'thumbs.db' in self.masslocal[3]:
                        self.masslocal[3].remove('thumbs.db')
                case "5":
                    #self.changefromat(pathfolder)
                    self.masslocal[4] = os.listdir(pathfolder)
                    if 'thumbs.db' in self.masslocal[4]:
                        self.masslocal[4].remove('thumbs.db')
                    elif 'Thumbs.db' in self.masslocal[4]:
                        self.masslocal[4].remove('Thumbs.db')
                case _:
                    continue
        print(f"\tЛокальных файлов: {len(self.masslocal[0])}\t{len(self.masslocal[1])}\t{len(self.masslocal[2])}\t{len(self.masslocal[3])}\t{len(self.masslocal[4])}")

        for element in os.listdir(mainpath):
            for elem in element:
                if elem == "'thumbs.db'":
                    print("\t\tЭлемент 'thumbs.db' найден!")

        print(f"Заканчиваем сканирование данных из локальных папок")

        # Если время для продвинутого сканирования, запускаем
        '''if self.date == times.timetoScan_2_0:
            print(f"Начинаем сканирование данных времени из локальных папок")
            # Пробегаемся по сформированному массиву, чтобы извлечь данные времени создания
            for element in tqdm(self.masslocal):
                # Вычисляем номер папки
                numberfolder = self.masslocal.index(element) + 1
                # Запускам цикл по фотографиям, которые находятся в папках
                for elem in element:
                    # Формируем путь к элементу
                    pathphoto = mainpath + str(numberfolder) + "/" + str(elem)
                    # Вычисляем размер изображения
                    datesize = os.stat(pathphoto).st_size
                    # Добавляем данные по результатам в массив
                    self.masslocalsize[numberfolder - 1].append(datesize)'''

    # Функция сканирования удалённых папкок с фотографиями
    def scanfilesinremoteserver(self):
        print(f"Начинаем сканирование данных из удалённых папок")
        # Инициализируем попытку сбора данных с удалённого сервера
        for element in os.listdir(server2path):
            # Формируем путь к папкам
            pathfolder = server2path + "/" + element
            #print(element, "\t\t", pathfolder)
            match element:
                # Первая папка для синхронизации
                case "1":
                    self.massremote[0] = os.listdir(pathfolder)
                    #self.massremote[0].remove('Thumbs.db')
                # Вторая папка для синхронизации
                case "2":
                    self.massremote[1] = os.listdir(pathfolder)
                    #self.massremote[1].remove('Thumbs.db')
                # Третья папка для синхронизации
                case "3":
                    self.massremote[2] = os.listdir(pathfolder)
                    #self.massremote[2].remove('Thumbs.db')
                # Четвёртая папка для синхронизации
                case "4":
                    self.massremote[3] = os.listdir(pathfolder)
                    #self.massremote[3].remove('Thumbs.db')
                # Пятая папка для синхронизации
                case "5":
                    self.massremote[4] = os.listdir(pathfolder)
                    #self.massremote[4].remove('Thumbs.db')
                case _:
                    continue

        print(f"Удалённых файлов: {len(self.massremote[0])}\t{len(self.massremote[1])}\t{len(self.massremote[2])}\t{len(self.massremote[3])}\t{len(self.massremote[4])}")

        for element in os.listdir(server2path):
            for elem in element:
                if elem == "'thumbs.db'":
                    print("\t\tЭлемент 'thumbs.db' найден!")

        print(f"Заканчиваем сканирование данных из удалённых папок")
        # Если время для продвинутого сканирования, запускаем
        '''if self.date == times.timetoScan_2_0:
            # Пробегаемся по сформированному массиву, чтобы извлечь данные времени создания
            for element in tqdm(self.massremote):
                # Открываем связь с удалённым сервером
                datesftp = FTP(ftpdates.nameSite)
                datesftp.login(ftpdates.ftpLogin, ftpdates.ftpPass)
                datesftp.set_pasv(False)
                # Вычисляем номер папки
                numberfolder = self.massremote.index(element) + 1
                # Запускам цикл по фотографиям, которые находятся в папках
                for elem in element:
                    # Формируем путь к элементу
                    remotepathphoto = "/" + str(numberfolder) + "/" + str(elem)
                    self.importremotedatesfromftp(datesftp, remotepathphoto, numberfolder)
                # Закрываем соединение с удалённым сервером
                datesftp.close()
        '''

    # Функция импорта данных
    def importatesfromftp(self, datesftp, element):
        # Определяем путь для папки
        path = "/" + str(element) + "/"
        # Изменение каталог работы
        datesftp.cwd(path)
        # Получаем лист всех фаилов из папки
        list = datesftp.nlst()
        # Удалем первые 2 элемента (так как на сервере система Linux)
        list.pop(0)
        list.pop(0)
        # Сортируем фаилы по возрастанию
        list.sort()
        # Возвращаем полученный список
        return list

    # Функция импорта данных дат файлов на удалённом сервере
    def importremotedatesfromftp(self, datesftp, remotepathphoto, numberfolder):
        sizephoto = datesftp.size(remotepathphoto)
        self.massremotesizes[numberfolder - 1].append(sizephoto)

    # Функция различия локальной у удалённой папки
    def comparisonlists(self):
        print(f"Начинаем выборку разницы локальных(215) и удалённых(35) папок")
        for element in range(0, 5):
            result = list(set(self.masslocal[element]) - set(self.massremote[element]))
            #print("Разница папки\t", element, "\t", len(result), "\t", result)
            if result == []:
                match element:
                    case 0:
                        print("Первые\t\tпапки синхронизированы!")
                    case 1:
                        print("Вторые\t\tпапки синхронизированы!")
                    case 2:
                        print("Третьи\t\tпапки синхронизированы!")
                    case 3:
                        print("Четвёртые\tпапки синхронизированы!")
                    case 4:
                        print("Пятые\t\tпапки синхронизированы!")
            else:
                match element:
                    case 0:
                        print("Разность первых папок: ", result)
                        self.uploadfiles(element + 1, result)
                    case 1:
                        print("Разность вторых папок: ", result)
                        self.uploadfiles(element + 1, result)
                    case 2:
                        print("Разность третьих папок: ", result)
                        self.uploadfiles(element + 1, result)
                    case 3:
                        print("Разность четвёртых папок: ", result)
                        self.uploadfiles(element + 1, result)
                    case 4:
                        print("Разность пятых папок: ", result)
                        self.uploadfiles(element + 1, result)

        print(f"Заканчиваем выборку разницы локальных(215) и удалённых(35) папок")

        # Если время для продвинутого сканирования, запускаем
        '''if self.date == times.timetoScan_2_0:
            print(f"Index\tElement\t\t\tNameLocal\tLocalSize\tRemoteName\t\tRemoteDate")
            for element in self.masslocal:
                print("===============================")
                indexfolder = self.masslocal.index(element)
                for i in element:
                    if element.index(i) == 10:
                        break
                    #else:
                        #print(f"{element.index(i)}\t\t{i}\t\t{self.masslocal[indexfolder][element.index(i)]}\t{self.masslocalsize[indexfolder][element.index(i)]}\t\t{self.massremote[indexfolder][element.index(i)]}")

                        #print(f"{element.index(i)}\t\t\t{i}\t\t{self.masslocal[indexfolder][element.index(i)]}\t{self.masslocalsize[indexfolder][element.index(i)]}\t\t{self.massremote[indexfolder][element.index(i)]}\t\t{self.massremotesizes[indexfolder][element.index(i)]}")
            print(f"\t\t\t\tFirstFolder\t\tSecondFolder\tThirdFolder\t\tFourthFolder\tFifthFolder")
            print(f"LocalMass:\t\t{len(self.masslocal[0])}\t\t\t{len(self.masslocal[1])}\t\t\t{len(self.masslocal[2])}\t\t\t{len(self.masslocal[3])}\t\t\t{len(self.masslocal[4])}")
            print(f"LocalMassSize:\t{len(self.masslocalsize[0])}\t\t\t{len(self.masslocalsize[1])}\t\t\t{len(self.masslocalsize[2])}\t\t\t{len(self.masslocalsize[3])}\t\t\t{len(self.masslocalsize[4])}")
            print(f"RemoteMass:\t\t{len(self.massremote[0])}\t\t\t{len(self.massremote[1])}\t\t\t{len(self.massremote[2])}\t\t\t{len(self.massremote[3])}\t\t\t{len(self.massremote[4])}")
            print(f"RemoteMassSize:\t{len(self.massremotesizes[0])}\t\t\t{len(self.massremotesizes[1])}\t\t\t{len(self.massremotesizes[2])}\t\t\t{len(self.massremotesizes[3])}\t\t\t{len(self.massremotesizes[4])}")
            print("=====================================")
            print(f"Localmass - RemoteMass")
            for i in range(0, 5):
                difference = len(self.masslocal[i])-len(self.massremote[i])
                if difference < 0:
                    mass = list(set(self.massremote[i]) - set(self.masslocal[i]))
                    print(f"Remote-Local {i} :\t\t{difference}\t\t{mass}")
                elif difference == 0:
                    print(f"Синхронизация  {i} не требудется")
                else:
                    mass = list(set(self.masslocal[i]) - set(self.massremote[i]))
                    print(f"Local-Remote {i} :\t\t{difference}\t\t{mass}")
                    '''

    # Функция загрузки фотографий на сервер
    def uploadfiles(self, numberfolder, result):
        print("Функция загрузки фотографий на сервер2")
        print("=====================")
        #print(numberfolder)
        #print(result)
        #print(len(result))
        for element in tqdm(result):
            if element != "thumbs.db":
                fistfilepath = mainpath + str(numberfolder) + "/" + element
                #print("\t\tОткуда:", fistfilepath)
                secondfilepath = server2path + "/" + str(numberfolder) + "/" + element.lower()
                #print("\t\tКуда:", secondfilepath)
                try:
                    #print("Нужно скопировать файл: ", element)
                    shutil.copyfile(fistfilepath, secondfilepath)
                except Exception:
                    print(f"Не удалить скопировать файл: {Exception}")
        print("=====================")

# Класс работы с Call центре
class class_call_center(object):

    def __init__(self, argument):
        # Приём веремени в класс
        self.date = argument
        # Приём данных по сотрудникам
        self.sotrud = allsotr
        # Приём данных по файлам excel
        self.datesexcel = datesforexcelfiles

    # Функция записи об обновлении файла Call центра
    def createnewarrowincallcenter2(self):
        try:
            # Подключаемся к сервисному аккаунту
            gc = gspread.service_account(CREDENTIALS_FILE)
            # Подключаемся к таблице по ключу таблицы
            table = gc.open_by_key(sheetkey)
            # Открываем нужный лист
            worksheet = table.worksheet("LogsCallCenter")
            # Получаем номер самой последней строки
            newstr = len(worksheet.col_values(1)) + 1
            # Вычисляем номер строки
            newnumber = newstr - 1
            # Определяем время выполения операции
            today = datetime.datetime.today().strftime("%d.%m.%Y | %H:%M:%S")
            # Определяем диапазон для обьединения ячеек
            mergerange = "C" + str(newstr) + ":F" + str(newstr)
            # Обьединяем ячейки да записи
            worksheet.merge_cells(mergerange)
            # Добавляем запись в таблицу логгирования
            worksheet.update_cell(newstr, 1, newnumber)
            worksheet.update_cell(newstr, 2, today)
            worksheet.update_cell(newstr, 3, "Фаил [График 2023 ТЕСТ.xlsx] обновлён")
            # Окрашивание ячейки
            color = {"backgroundColor": {"red": 0.94, "green": 0.9, "blue": 0.15}, "horizontalAlignment": "CENTER"}
            worksheet.format("C" + str(newstr), color)
            # Делаем центрирование ячейки
            worksheet.format(mergerange, {"horizontalAlignment": "CENTER"})
        except Exception as e:
            print(f"Логгирование call-центра сломалось: {e}")
            time.sleep(5)
            self.createnewarrowincallcenter2()

    # Функция копирования данный в файл для работы
    def checkupdatedatesexcel(self):
        # Вычисляем время последнего изменения основного документа
        #file1 = openpyxl.load_workbook(mainfile).properties.modified
        file1 = openpyxl.load_workbook(datesforexcelfiles.pathmainfile).properties.modified
        # Вычисляем дату последнего изменения рабочего документа
        #file2 = openpyxl.load_workbook(pathfile).properties.modified
        file2 = openpyxl.load_workbook(datesforexcelfiles.pathfile).properties.modified

        # Вычисляем разницу времён
        diff_times = file1 - file2
        # Устанавливаем время для синхронизации фаилов
        deltatime = datetime.timedelta(days=0, hours=0, minutes=5)
        # Если deltatime меньше разницы во времени изменения файлов
        if deltatime < diff_times:
            print("\t\tСинхронизация требуется")
            # Выполняем копирование
            shutil.copy2(datesforexcelfiles.pathmainfile, datesforexcelfiles.pathfile, follow_symlinks=True)
            print("\t\tСтатус операции изменения фаила: [Фаил обновлён]")
            # Логгирование обновления фаила
            self.createnewarrowincallcenter2()
        # Иначе синхронизацию не выполняем.
        else:
            print("Синхронизация не требуется")

    # Функция импорта данных из файла для работы
    def importdatesformexcel(self, path, password):
        # Экземпляр COM обьекта
        xlApp = win32com.client.Dispatch("Excel.Application")
        # Открываем фаил
        xlwb = xlApp.Workbooks.Open(path, False, True, None, password)
        # Выбираем лист(таблицу)
        sheet = xlwb.ActiveSheet
        # Выбираем данные из range
        alldates = sheet.Range("B1:B451")

        # Выясняем текущий мясяц
        today = datetime.datetime.today()
        todayyear = int(today.strftime("%Y"))
        listmontheng = [datetime.date(todayyear, 1, 1).strftime("%B"), datetime.date(todayyear, 2, 1).strftime("%B"),
                        datetime.date(todayyear, 3, 1).strftime("%B"), datetime.date(todayyear, 4, 1).strftime("%B"),
                        datetime.date(todayyear, 5, 1).strftime("%B"), datetime.date(todayyear, 6, 1).strftime("%B"),
                        datetime.date(todayyear, 7, 1).strftime("%B"), datetime.date(todayyear, 8, 1).strftime("%B"),
                        datetime.date(todayyear, 9, 1).strftime("%B"), datetime.date(todayyear, 10, 1).strftime("%B"),
                        datetime.date(todayyear, 11, 1).strftime("%B"), datetime.date(todayyear, 12, 1).strftime("%B")]
        listmonthrus = ["ЯНВАРЬ", "ФЕВРАЛЬ", "МАРТ", "АПРЕЛЬ", "МАЙ", "ИЮНЬ", "ИЮЛЬ", "АВГУСТ", "СЕНТЯБРЬ", "ОКТЯБРЬ",
                        "НОЯБРЬ", "ДЕКАБРЬ"]
        todaymontheng = today.strftime("%B")
        todaymonthrus = listmonthrus[listmontheng.index(todaymontheng)]

        # Ищем стартовую ячейку, для определения графика на этот месяц
        index = 0
        indexmonth = 0
        # Перебираем все элементы и находим нужную ячейку с текущим месяцем
        for element in alldates:
            index = index + 1
            if str(element) == todaymonthrus:
                indexmonth = index
        # Формируем название ячейки начала импорта
        firstcell = "B" + str(indexmonth)
        # Формируем название ячейки конца импорта
        lastcell = "AG" + str(indexmonth + 31)
        # Формируем строку для импорта
        cellsrange = firstcell + ":" + lastcell
        # Импортируем данные за нужный нам месяц
        datesforsolution = sheet.Range(cellsrange)
        # Формируем данные в список
        listdatesforsolution = []
        for element in datesforsolution:
            listdatesforsolution.append(str(element))
        # Закрываем фаил
        xlwb.Close()
        # Закрываем COM обьект
        xlApp.Quit()

        # Возвращаем данные
        return listdatesforsolution

    # Функция для выбоки по данным
    def chosedates(self, dates):
        # Удаляем первый элемент
        del dates[0]

        # Считаем дни в месяце
        index = 0
        for element in dates:
            index = index + 1
            if element == "None" or element == "Торговля":
                # Определяем количество дней в месяце
                countdaysinmonth = index - 1
                break

        # Удаляем ненужные данные
        for element in dates:
            #if element == massmanagers[0]:
            if element == allsotr.massmanagers_short[0]:
                delelements = dates.index(element)
        del dates[0:delelements]

        # Разбиваем массив для конкретизации графика каждого менеджера
        managerlists = []
        #lenmanagers = len(massmanagers) - 1
        lenmanagers = len(allsotr.massmanagers_short) - 1
        for i in range(0, lenmanagers):
            managerlist = []
            index = 0
            for element in dates:
                managerlist.append(element)
                index += 1
                if index == 32:
                    break
            managerlists.append(managerlist)
            del dates[0:32]

        # Выясняем график работы ПП
        deldates = 32 * 8
        # Удаляем ненужные данные
        del dates[0:deldates]
        # Добавляем в массив работников данные
        managerlist = []
        index = 0
        for element in dates:
            managerlist.append(element)
            index = index + 1
            if index == countdaysinmonth + 1:
                break
        managerlists.append(managerlist)
        return managerlists

    # Функция активирования менеджеров
    def selectmenegers(self, managerlists):
        # Выясняем текущй день
        today = datetime.datetime.today()
        todayday = int(today.strftime("%d"))
        print("Сегодня:", todayday, today.strftime("%B"), int(today.strftime("%Y")))
        flag = True
        massworkmanagers = []
        try:
            # Изменяем статусы менеджеров call центра
            for element in managerlists:
                if element[todayday] == "В" or element[todayday] == "O" or element[todayday] == "О" or element[todayday] == "Х":
                    #numbermanager = numbermanagers[massmanagers.index(element[0])]
                    numbermanager =  allsotr.numbermanagers[allsotr.massmanagers_short.index(element[0])]
                    print("\t\tНеобходимо деактивировать телефон: ", element[0], "\t[", element[todayday], "]", "'",
                          numbermanager,
                          "'")
                    urlforapi = urlapi + str(numbermanager) + '/agent'
                    statusrequest = requests.put(urlforapi, params=paramoffline, headers=headers)
                    if statusrequest == "<Response [403]>":
                        flag = False
                        print("\tЧто-то пошло не так... Нет ответа по запросу изменения статуса")
                    else:
                        statusget = requests.get(urlforapi, headers=headers).text
                        print("\tСтатус менеджера: ", element[0], " = ", statusget)
                else:
                    #numbermanager = numbermanagers[massmanagers.index(element[0])]
                    numbermanager = allsotr.numbermanagers[allsotr.massmanagers_short.index(element[0])]
                    print("\t\tНеобходимо активировать телефон: ", element[0], "\t[", element[todayday], "]", "'",
                          numbermanager,
                          "'")
                    urlforapi = urlapi + str(numbermanager) + '/agent'
                    statusrequest = requests.put(urlforapi, params=paramsonline, headers=headers)
                    if statusrequest == "<Response [403]>":
                        flag = False
                        print("\tЧто-то пошло не так... Нет ответа по запросу изменения статуса")
                    else:
                        # Дополнительное условие для последнего менеджера
                        massworkmanagers.append(element[todayday])
                        if len(massworkmanagers) == 4:
                            # Если 3 других менеджера работают, то 4 должен быть отключён
                            if (massworkmanagers[0] == '9.0' or massworkmanagers[0] == '10.0') and (
                                    massworkmanagers[1] == '9.0' or massworkmanagers[1] == '10.0') and (
                                    massworkmanagers[2] == '9.0' or massworkmanagers[2] == '10.0'):
                                requests.put(urlforapi, params=paramoffline, headers=headers)
                        statusget = requests.get(urlforapi, headers=headers).text
                        print("\tСтатус менеджера: ", element[0], " = ", statusget)

            if flag == True:
                return "\tCall центр успешно настроен."
            else:
                return "\tВ работе функции произошла ошибка"
        except Exception as e:
            print(f"В работе call-центра произошла ошибка: {e}")

    # Функция записи логов Call Center
    def createnewarrowincallcenter(self):
        try:
            # Подключаемся к сервисному аккаунту
            gc = gspread.service_account(CREDENTIALS_FILE)
            # Подключаемся к таблице по ключу таблицы
            table = gc.open_by_key(sheetkey)
            # Открываем нужный лист
            worksheet = table.worksheet("LogsCallCenter")
            # Получаем номер самой последней строки
            newstr = len(worksheet.col_values(1)) + 1
            # Вычисляем номер строки
            newnumber = newstr - 1
            # Определяем время выполения операции
            today = datetime.datetime.today().strftime("%d.%m.%Y | %H:%M:%S")
            # Выясняем данные кто работает
            managerslist = []
            # Выясняем статусы менеджеров
            #for element in numbermanagers:
            for element in allsotr.numbermanagers:
                urlforapi = urlapi + element + '/agent'
                status = requests.get(urlforapi, headers=headers).text
                managerslist.append(status)
            # Проверяем изменится ли call центр
            dates = worksheet.row_values(newnumber)
            # Если данные уже сегодня записывались, то не дублируем их
            if dates[2] == managerslist[0] and dates[3] == managerslist[1] and dates[4] == managerslist[2] and dates[
                5] == managerslist[3] and str(dates[1])[:10] == str(today)[:10]:
                print("\t\tДанные уже были записаны")
            # Если же эти данные не были записаны, записываем
            else:

                # Добавляем строку в конец фаила логгирования
                worksheet.update_cell(newstr, 1, newnumber)
                worksheet.update_cell(newstr, 2, today)

                for element in range(0, 4):
                    if managerslist[element] == '"ONLINE"' or managerslist[element] == '"OFFLINE"':
                        worksheet.update_cell(newstr, element + 3, managerslist[element])
                    else:
                        worksheet.update_cell(newstr, element + 3, "Ошибка данных")

                if managerslist[0] == '"ONLINE"':
                    worksheet.format("C" + str(newstr), colorsforbuttons.greencolor)
                else:
                    worksheet.format("C" + str(newstr), colorsforbuttons.redcolor)

                if managerslist[1] == '"ONLINE"':
                    worksheet.format("D" + str(newstr), colorsforbuttons.greencolor)
                else:
                    worksheet.format("D" + str(newstr), colorsforbuttons.redcolor)
                if managerslist[2] == '"ONLINE"':
                    worksheet.format("E" + str(newstr), colorsforbuttons.greencolor)
                else:
                    worksheet.format("E" + str(newstr), colorsforbuttons.redcolor)
                if managerslist[3] == '"ONLINE"':
                    worksheet.format("F" + str(newstr), colorsforbuttons.greencolor)
                else:
                    worksheet.format("F" + str(newstr), colorsforbuttons.redcolor)
                # Чтобы программа не крашилась из-за лимита количества запросов ставим sleep
                datetime.time.sleep(60)
        except Exception as e:
            print(f"Логгирование call-центра сломалось: {e}")
            time.sleep(10)
            self.createnewarrowincallcenter()

    # Функция изменения Call центра
    def changecallcenter(self):
        # Инициализация многопоточности
        pythoncom.CoInitialize()
        # Функция проверки файла на актуальность
        self.checkupdatedatesexcel()
        # Достаём данные из файла
        datesnowmonth = self.importdatesformexcel(datesforexcelfiles.pathfile, datesforexcelfiles.password)
        # Выбираем данные для работы с ними
        massive = self.chosedates(datesnowmonth)
        # Активируем телефоны менеджеров
        result = self.selectmenegers(massive)
        # Записываем изменения в таблицу логгирования
        self.createnewarrowincallcenter()

# Класс работы со сбором статистики по звонкам
class class_collecion_of_information(object):

    # Счётчик попыток
    attemptcounter = 0
    # Массив звонков
    calls = []
    # Массив данных
    dates = []
    # Массив для подсчёта пропущенных звонков
    massmissescals = [0, 0, 0, 0]
    # Массив для подсчёта принятых звонков
    massinboundcalls = [0, 0, 0, 0]
    # Массив для подсчёта общего времени общения с клиентами
    masssumtimes = [datetime.timedelta(milliseconds=0), datetime.timedelta(milliseconds=0),
                    datetime.timedelta(milliseconds=0), datetime.timedelta(milliseconds=0)]
    # Массив менеджеров которые сегодня работали
    workedmanagers = [0, 0, 0]
    # Массив названий столбцов
    masscolumns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
    # Дата начала отчёта (вчерашний день начало для)
    dateAndTimeStart = ""
    # Дата окончания отчёта (сегодняшний день начало дня)
    dateAndTimeEnd = ""

    # Инициализация данных переменных
    def __init__(self, argument):
        self.attemptcounter = 0
        self.calls = []
        self.dates = []
        self.massmissescals = [0, 0, 0, 0]
        self.massinboundcalls = [0, 0, 0, 0]
        self.masssumtimes = [datetime.timedelta(milliseconds=0), datetime.timedelta(milliseconds=0), datetime.timedelta(milliseconds=0), datetime.timedelta(milliseconds=0)]
        self.workedmanagers = [0, 0, 0]
        self.masscolumns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
        self.dateAndTimeStart = ""
        self.dateAndTimeEnd = ""

        self.date = argument
        self.dateAndTimeStart = str((datetime.datetime.today() + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")) + "T00:00:00.000Z"
        self.dateAndTimeEnd = str(datetime.datetime.today().strftime("%Y-%m-%d")) + "T00:00:00.000Z"

    # Функциия импорта и систематизация статистики по звонкам
    def collectionofinformation(self):

        # Класс звонка
        class phoneCall:
            def __init__(self, name_manager, incoming_call_time, incoming_call_number, call_duration, direction, status):
                # ФИО менеджера
                self.name_manager = name_manager
                # Время входящего звонка
                self.incoming_call_time = incoming_call_time
                # Телефон звонка
                self.incoming_call_number = incoming_call_number
                # Продолжительность звонка
                self.call_duration = call_duration
                # Тип вызова (INBOUND-Входящий вызов, OUTBOUND-Исходящий вызов))
                self.direction = direction
                # Статус звонка (RECIEVED-принятый, MISSED-пропущенный, PLACED-исходящий)
                self.status = status

            # Функция печати данных
            def printdates(self):
                print(f"\t{self.name_manager}"
                      f"\t{self.incoming_call_time}"
                      f"\t\t{self.incoming_call_number}"
                      f"\t{self.call_duration}"
                      f"\t\t{self.direction}"
                      f"\t\t{self.status}")

        try:
            # Пробегаемся по списку менеджеров
            #for element in numbermanagers:
            for element in allsotr.numbermanagers:
                # Формируем данные для запроса
                paramsinfo['userId'] = element
                paramsinfo['dateTo'] = self.dateAndTimeEnd
                paramsinfo['dateFrom'] = self.dateAndTimeStart
                # Делаем запрос к API
                statusrequest = requests.get(urlforstatistics, params=paramsinfo, headers=headers)
                # Формираем Json
                jsonData = json.loads(statusrequest.text)
                # Разбираем данные
                for elem in jsonData:
                    # Вычисляем время звонка
                    dateandtime = datetime.datetime.fromtimestamp(elem['startDate'] / 1000)
                    # Вычисляем продолжительность разговора
                    dateandtime2 = datetime.timedelta(milliseconds=elem['duration'])
                    # Если вызов входящий
                    if elem['direction'] == 'INBOUND':
                        # Вычисляем телефон абонента
                        phone = elem['phone_from']
                    # Если вызов исходящий:
                    elif elem['direction'] == 'OUTBOUND':
                        # Вычисляем телефон абонента
                        phone = elem['phone_to']
                    else:
                        continue
                    # Добавляем в массив звонков экземпляр класса phoneCall
                    self.addElemInMassCalls(phoneCall, elem, dateandtime, phone, dateandtime2)
                # Заполняем массив статусов
                urlforapi = urlapi + element + '/agent'
                status = requests.get(urlforapi, headers=headers).text
                for elem in range(0, 3):
                    self.workedmanagers.append(status)

            # Функция добавления данных
            self.addDates()
            # Функция сортировки данных по звонкам
            self.sortCalls()
            # Данные по таблице
            worksheet = self.datesGoogleTable()
            # Проверяем были ли записаны данные ранее
            datesfromtabel = worksheet.row_values(len(worksheet.col_values(4)))
            # Записываем получившееся результаты в таблицу
            #self.InsertDatesInTable()

            # Дата начала отчёта (вчерашний день начало для)
            print(f"\t\tdateAndTimeStart = {self.dateAndTimeStart}")

            # Дата окончания отчёта (сегодняшний день начало дня)
            print(f"\t\tdateAndTimeEnd = {self.dateAndTimeEnd}")

            # Записываем получившееся результаты в таблицу
            if datesfromtabel[2] == self.dates[2]:
                print(f"\t\tДанные [{datesfromtabel[2]} = {self.dates[2]}] уже были записаны")
            else:
                self.InsertDatesInTable()

        except Exception as e:
            self.attemptcounter += 1
            print(f"Попытка: {self.attemptcounter}\nЛоггирование статистики по звонкам сломалось: {e}")

            time.sleep(20)
            if self.attemptcounter <= 10:
                self.collectionofinformation()
            else:
                print(f"Попытка: {self.attemptcounter} закончилась неудачно.")

    # Функция разбора данных по звонкам
    def addinfoinmass(self, massmissescals, massinboundcalls, masssumtimes, numbermanager, elemclass):
        # Если вызов входящий пропущенный
        if elemclass.direction == "INBOUND" and elemclass.status == "MISSED":
            massmissescals[numbermanager] += 1
        # Если вызов входящий принятый
        elif elemclass.direction == "INBOUND" and elemclass.status == "RECIEVED":
            massinboundcalls[numbermanager] += 1
            masssumtimes[numbermanager] += elemclass.call_duration
        return [massmissescals, massinboundcalls, masssumtimes]

    # Функция удобно представления времени разговора из миллисекунд в нормальное представление
    def converttoseconds(self, totseconds):
        hours, remainder = divmod(int(totseconds), 3600)
        minutes, seconds = divmod(remainder, 60)
        result = str(hours) + ":" + str(minutes) + ":" + str(seconds)
        return result

    # Функция добавления в массив звонков
    def addElemInMassCalls(self, phoneCall, jsonelem, dateandtime, phone, dateandtime2):
        self.calls.append(phoneCall(name_manager = jsonelem['abonent']['firstName'],
                                    incoming_call_time=dateandtime,
                                    incoming_call_number=phone,
                                    call_duration=dateandtime2,
                                    direction=jsonelem['direction'],
                                    status=jsonelem['status']))

    # Данные по таблице Google Sheets
    def datesGoogleTable(self):
        # Подключаемся к сервисному аккаунту
        gc = gspread.service_account(CREDENTIALS_FILE)
        # Подключаемся к таблице по ключу таблицы
        table = gc.open_by_key(sheetkey)
        # Открываем нужный лист
        worksheet = table.worksheet("StatisticOfCalls")
        return worksheet

    # Функция добавление данных в массив данных
    def addDates(self):
        # Данные по таблице
        worksheet = self.datesGoogleTable()
        # Получаем номер самой последней строки
        newstr = len(worksheet.col_values(4)) + 1
        # Вычисляем номер строки
        newnumber = newstr - 2
        # Добавляем номер строки
        self.dates.append(newnumber)
        # Определяем время выполения операции
        today = datetime.datetime.today().strftime("%d.%m.%Y | %H:%M:%S")
        # Добавляем дату выполения операции
        self.dates.append(today)
        # Вычисляем дату за которую приводим статистику
        statdate = (datetime.datetime.today() + datetime.timedelta(days=-1)).strftime("%d.%m.%Y")
        # Добавляем дату за которую приводим статистику
        self.dates.append(statdate)

    # Функция сортировки данных по звонкам
    def sortCalls(self):
        # Пробегаемся по всем звонкам и сортируем звонки
        for element in self.calls:
            # Считаем статистику для первого менеджера
            #if element.name_manager == fullmassmanagers[0]:
            if element.name_manager == allsotr.fullmassmanagers[0]:
                self.addinfoinmass(self.massmissescals, self.massinboundcalls, self.masssumtimes, 0, element)
            # Считаем статистику для второго менеджера
            #elif element.name_manager == fullmassmanagers[1]:
            elif element.name_manager == allsotr.fullmassmanagers[1]:
                self.addinfoinmass(self.massmissescals, self.massinboundcalls, self.masssumtimes, 1, element)
            # Считаем статистику для третьего менеджера
            #elif element.name_manager == fullmassmanagers[2]:
            elif element.name_manager == allsotr.fullmassmanagers[2]:
                self.addinfoinmass(self.massmissescals, self.massinboundcalls, self.masssumtimes, 2, element)
            # Считаем статистику для четвёртого менеджера
            #elif element.name_manager == fullmassmanagers[3]:
            elif element.name_manager == allsotr.fullmassmanagers[3]:
                self.addinfoinmass(self.massmissescals, self.massinboundcalls, self.masssumtimes, 3, element)
            else:
                print("Cтатистика для Неизвестного лица(")
        # Добавляем данные с разбора в результурующий массив
        for element in range(4):
            self.dates.append(self.massmissescals[element])
            self.dates.append(self.massinboundcalls[element])
            self.dates.append(self.converttoseconds(self.masssumtimes[element].total_seconds()))

    # Функция вставки данных в таблицу
    def InsertDatesInTable(self):
        # Данные по таблице
        worksheet = self.datesGoogleTable()
        # Получаем номер самой последней строки
        newstr = len(worksheet.col_values(4)) + 1
        i = 0
        for element in self.dates:
            match self.workedmanagers[i]:
                case '"ONLINE"':
                    worksheet.update_cell(newstr, i + 1, self.dates[i])
                    worksheet.format(self.masscolumns[i] + str(newstr), colorsforworkers.colorwork)
                case '"OFFLINE"':
                    worksheet.update_cell(newstr, i + 1, self.dates[i])
                    worksheet.format(self.masscolumns[i] + str(newstr), colorsforworkers.coloroutput)
                case _:
                    worksheet.update_cell(newstr, i + 1, self.dates[i])
                    worksheet.format(self.masscolumns[i] + str(newstr), colorsforworkers.colornone)
            i += 1

# Класс сохранения статистики по загруженным фотографиям
class class_generation_stat_uploadphotos(object):

    # Массив значений
    massvalues = []
    # Массив столбцов для импорта статистики
    masscolumns = ["L", "M", "N", "O", "P", "Q", "R"]

    def __init__(self, argument):
        self.time = argument

    # Функция сохранения статистики по загруженным фотографиям
    def generationstatuploadphotos(self):
        try:
            # Проверяем дату сегодняшнюю
            today = datetime.datetime.today()
            todaytime = today.strftime("%d")
            print(f"\tДата сейчас: {todaytime}")

            # Вычисляем день недели
            day_of_week = today.weekday()
            print(f"День недели: {day_of_week}")

            # Если сегодня суббота
            if day_of_week == 5:
                # Вычисляем период за который сохраняем статистику
                startdate = (datetime.datetime.today() - datetime.timedelta(days=7)).strftime("%d.%m.%Y")
                enddate = datetime.datetime.today().strftime("%d.%m.%Y")
                resultdates = str(startdate) + " : " + str(enddate)
                print(f"{startdate} - {enddate} : {resultdates}")

                # Подключаемся к сервисному аккаунту
                gc = gspread.service_account(CREDENTIALS_FILE)
                # Подключаемся к таблице по ключу таблицы
                table = gc.open_by_key(sheetkey)
                # Открываем нужный лист
                worksheet = table.worksheet("LogsPhotos")
                # Получаем номер строки для записи в столбце L
                newstr = len(worksheet.col_values(12)) + 1
                # Получаем данные из столбца H
                massvalues = worksheet.get_values('H2:H6')
                sumphotos = 0

                # Преобразовываем массив
                for element in massvalues:
                    self.massvalues.append(int(element[0]))
                    sumphotos += int(element[0])

                print(massvalues)
                print(sumphotos)

                # Запись данных в табличку
                for element in range(0, 7):
                    column = element + 12
                    print(column)
                    match column:
                        case 12:
                            print(f"Update 12 | {newstr} | {column} | {resultdates}")
                            worksheet.update_cell(newstr, column, resultdates)
                            worksheet.format(self.masscolumns[element] + str(newstr), colorsforworkers.colornone)
                        case 18:
                            print(f"Update 18 | {newstr} | {column} | {sumphotos}")
                            worksheet.update_cell(newstr, column, sumphotos)
                            worksheet.format(self.masscolumns[element] + str(newstr), colorsforworkers.colornone)
                        case dafault:
                            print(f"Update default | {newstr} | {column} | {self.massvalues[element - 1]}")
                            worksheet.update_cell(newstr, column, self.massvalues[element - 1])
                            worksheet.format(self.masscolumns[element] + str(newstr), colorsforworkers.colornone)
                # Обнуляем значения, которые подсчитываются онлайн
                for element in range(2, 7):
                    worksheet.update_cell(element, 8, 0)
                # Записываем дату с коротой считаются фотографии
                nulldate = today.strftime("%d %m %Y")
                worksheet.update_cell(2, 9, nulldate)
                text = f"\t Подсчёт загруженных фотографий навершён"
                print(text)
                return text
            # Иначе ничего не делаем
            else:
                text = f"\tВремя для обнуления ещё не пришло."
                print(text)
                return text
        except Exception as e:
            text = f"Логгирование статистики фотографий сломалось: {e}"
            print(text)

# Класс отправки сообщений от телеграмм бота
class class_send_erorr_message(object):
    # Инициализация класса
    def __init__(self, argument, text, exception, botkey):
        self.time = argument
        self.function = text
        self.exception = exception
        self.botkey = botkey

    # Функция отправки сообщения об ошибке администратору, системному администратору
    def send_message(self):
        # Формирование сообщения
        message = "Возникла проблема с функцией: " + str(self.function) + " [" + str(self.time) + "]\n" + "Ошибка типа:\n{" + str(self.exception) + "}\n"
        # Токен для связи с ботом
        bot = telebot.TeleBot(botkey)
        # Отравляем сообщение на рабочий телефон администратора
        bot.send_message(1871580124,text=message)
        # Отравляем сообщение на личный телефон системного администратора
        bot.send_message(1917167694, text=message)
        return message

# Класс проверки файлов на актуальность
class class_checks(object):

    def __init__(self, argument):
        self.time = argument

    def start(self):
        # Проверяем когда был изменён файл базы данных
        pathtodatabase = '//SPEED-DEMON-II/wwwroot/files/database.db'
        self.decisionmaking(pathtodatabase)
        # Проверяем когда был изменён файл Дром
        pathtodrom = '//SPEED-DEMON-II/wwwroot/files/dromoutputinstock.xml'
        self.decisionmaking(pathtodrom)
        # Проверяем когда был изменён файл Авито
        pathtoavito = '//SPEED-DEMON-II/wwwroot/files/avitooutputinstock.xml'
        self.decisionmaking(pathtoavito)
        # Проверяем когда был изменён файл Авито
        doublegis = '//SPEED-DEMON-II/wwwroot/files/2gisoutputinstock.yml'
        self.decisionmaking(doublegis)

        # Функция выяснения последней модификации файла

    # Функция получения последней модификации файла
    def getlastmodifieddate(self, file_path):
            try:
                timestamp = os.stat(file_path).st_mtime
                return datetime.datetime.fromtimestamp(timestamp)
            except FileNotFoundError:
                return None

    # Функция отправки сообщения и принятия решения
    def decisionmaking(self, pathfile):
        # Получаем сегодняшнюю дату
        today = datetime.datetime.today()
        # Максимальное время без обновления
        oneday = datetime.timedelta(days=1)
        # Получаем дату последнего изменения файла
        lasttimeupdate = self.getlastmodifieddate(pathfile)
        print("Файл: ", pathfile)
        print("\tДата изменения файла : ", lasttimeupdate.strftime("%d.%m.%Y %H:%M:%S"))
        difference = today - lasttimeupdate
        print("\tВремя с последнего обновления: ", difference)
        if difference > oneday:
            print("\tНеобходимо обновить файл.")
            text = "\n\nФайл: \n" + pathfile + "\nНе обновился!\n" + str(difference) + "\n"
            exception = "NoException"
            # Инициализация класса
            error_message = class_send_erorr_message(today.strftime("%H:%M"),
                                                     text,
                                                     exception,
                                                     botkey)
            # Функция отправки сообщения в чат системному администратору
            error_message.send_message()
        else:
            print("\tОбновление файла не требуется")